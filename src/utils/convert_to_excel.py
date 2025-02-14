import logging
import openpyxl

from src.repositories.channel import ChannelRepository
from src.repositories.post import PostRepository
from src.repositories.comment import CommentRepository

file_name = 'output/output.xlsx'

def create_or_load_workbook():
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Пост айди', 'Тип канала', 'Ссылка на канал', 'Пост', 'Комментарии', 'Пользователь', 'Реакции Позитивные', 'Реакции Негативные', 'dt комментария'])
    return workbook

def collect_data(post, channel_type, channel_id, comment):
    post_id = post.post_id
    post_text = post.text if post.text is not None else post.media
    user_id = comment.user_id
    reactions_positive = post.reactions.split(':')[0]
    reactions_negative = post.reactions.split(':')[1]
    dt = int((comment.time - post.time).total_seconds() / 60)
    
    return [post_id, channel_type, channel_id, post_text, comment.text, user_id, reactions_positive, reactions_negative, dt]

async def converting(
        channel_repo: ChannelRepository,
        post_repo: PostRepository,
        comment_repo: CommentRepository,
):
    workbook = create_or_load_workbook()
    worksheet = workbook.active
    for post in post_repo.get_posts():
        try:
            channel_type = channel_repo.get_channel_type_by_id(
                    f"https://t.me/{post.url.split('/')[1]}"
                    )
        except Exception as e:
            channel_type = 'неизвестный'
        
        for comment in comment_repo.get_comments_by_post_id(post.post_id):
            logging.info(f'{comment.id} - {post.post_id}')
            data = collect_data(post, channel_type, f"https://t.me/{post.url.split('/')[1]}", comment)
            worksheet.append(data)

        workbook.save(filename=file_name)
