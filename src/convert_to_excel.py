import logging
import openpyxl
from .db.connection import Connection
from .db.sync_crud import get_channel_type_by_id, get_posts, get_comments_by_post_id

file_name = 'output.xlsx'

def create_or_load_workbook():
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(['Пост айди', 'Тип канала', 'Ссылка на канал', 'Пост', 'Комментарии', 'Пользователь', 'Реакции Позитивные', 'Реакции Негативные', 'dt комментария'])
    return workbook

def collect_data(post, channel_type, channel_id, comment):
    post_id = post.post_id
    post_text = post.text if post.text is not None else post.media
    user_id = comment.user_id
    reactions_positive = post.reactions.split(':')[0]
    reactions_negative = post.reactions.split(':')[1]
    dt = int((comment.time - post.time).total_seconds() / 60)
    
    return [post_id, channel_type, channel_id, post_text, comment.text, user_id, reactions_positive, reactions_negative, dt]

async def converting():
    async with Connection.getConnection() as session:
        workbook = create_or_load_workbook()
        worksheet = workbook.active
        for post in get_posts(session=session):
            try:
                channel_type = get_channel_type_by_id(
                        session, 
                        f"https://t.me/{post.url.split('/')[1]}")
            except Exception as e:
                channel_type = 'неизвестный'
            
            for comment in get_comments_by_post_id(session, post.post_id):
                logging.info(f'{comment.id} - {post.post_id}')
                data = collect_data(post, channel_type, f"https://t.me/{post.url.split('/')[1]}", comment)
                worksheet.append(data)

        workbook.save(filename=file_name)
